from django.shortcuts import get_object_or_404, render
from .models import Category, Product
import pandas as pd
from django.http import JsonResponse
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils import timezone
from openpyxl import load_workbook
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count
from slugify import slugify
import decimal
from decimal import Decimal
from django.db import IntegrityError

def import_data(request):
    if request.method == 'POST':
        # Отримуємо файл з запиту POST
        file = request.FILES.get('file')
        if not file:
            return render(request, 'import.html', {'error': 'Файл не було вибрано'})

        # Відкриваємо файл з допомогою openpyxl
        workbook = load_workbook(file)
        # Отримуємо потрібний аркуш
        worksheet = workbook.active

        import_success = True  # допоміжна змінна для відстеження успіху імпорту
        # Створити порожній список продуктів
        products_to_create = []

        try:
            # Отримуємо дані з кожного рядка у файлі
            for row in worksheet.iter_rows(min_row=2):
                id = row[7].value
                if not str(id).isdigit():
                    import_success = False
                    return render(request, 'import.html', {'error': f'Поле "id" має некоректне значення: "{id}"'})
                id = int(id)
                if Product.objects.filter(id=id).exists():
                    import_success = False
                    return render(request, 'import.html', {'error': f'Продукт з id {id} вже існує'})
                brand = row[0].value
                category_id = row[2].value
                created = row[3].value
                created_by_id = row[5].value
                description = row[6].value
                image = row[8].value
                in_stock = row[9].value
                is_active = row[10].value
                slug = row[12].value or slugify(row[13].value)
                if Product.objects.filter(slug=slug).exists():
                    import_success = False
                    return render(request, 'import.html', {'error': f'Продукт з slug {slug} вже існує'})
                title = row[13].value
                category_name = row[1].value
                updated = row[14].value
                price_str = row[11].value
                
                if not price_str:
                    import_success = False
                    return render(request, 'import.html', {'error': 'Поле "price" не може бути порожнім'})

                try:
                    price = Decimal(price_str)
                except (decimal.InvalidOperation, TypeError, ValueError):
                    import_success = False
                    return render(request, 'import.html', {'error': 'Ви ввели некоректну ціну'})
                
                price_str = str(row[11].value)
                
                if ',' in price_str:
                    if not price_str.replace(',', '').isnumeric():
                        import_success = False
                        return render(request, 'import.html', {'error': 'Ви ввели некоректну ціну'})
                    price_str = price_str.replace(',', '.')

                if not category_id or not category_name:
                    import_success = False
                    return render(request, 'import.html', {'error': 'Не вказано категорію для продукту'})

                if category_id:
                    try:
                        category = Category.objects.get(id=category_id)
                    except Category.DoesNotExist:
                        category = Category(id=category_id, name=category_name)
                        category.save()
                else:
                    try:
                        category = Category.objects.get(name=category_name)
                    except Category.DoesNotExist:
                        category = Category(name=category_name, slug=slugify(category_name))
                        category.save()

                if category.id != category_id or category_name != category.name:
                    import_success = False
                    return render(request, 'import.html', {'error': f'Категорія з id {category_id} вже існує з іншим ім\'ям'})
                
                if import_success == False:
                    raise Exception('Помилка імпортування')
                if import_success == True:
                    # Створюємо новий об'єкт Product з отриманими даними і зберігаємо його
                    product = Product(
                        brand=brand,
                        category=category,
                        created=created,
                        created_by_id=created_by_id,
                        description=description,
                        id=id,
                        image=image,
                        in_stock=in_stock,
                        is_active=is_active,
                        price=price,
                        slug=slug or slugify(title),  # якщо slug не задано, генеруємо його з title
                        title=title,
                        updated=updated
                    )
                    # Додати продукт до списку
                    products_to_create.append(product)
        
        except Exception as e:
            # обробка помилки
            print('Сталась помилка:', str(e))
            return render(request, 'import.html', {'error': 'Сталась помилка під час імпортування даних'})
        
        slugs = [product.slug for product in products_to_create]
        if len(set(slugs)) != len(slugs):
            import_success = False
            return render(request, 'import.html', {'error': 'Є продукти з однаковими slug'})
        
        
        # Зберегти всі продукти одночасно
        try:
            Product.objects.bulk_create(products_to_create)
        except IntegrityError as e:
            print('Помилка збереження продуктів: ', e)         

        return render(request, 'import.html', {'success': 'Дані успішно імпортовано!'})
    else:
        return render(request, 'import.html')

def export_excel(request):
    # Отримати дані з моделі Product
    products = Product.objects.all().values('brand', 'category', 'category_id', 'created', 'created_by', 'created_by_id', 'description', 'id', 'image', 'in_stock', 'is_active', 'order_items', 'price', 'slug', 'title', 'updated')

    # Створити новий файл Excel
    wb = Workbook()
    ws = wb.active

    # Додати заголовки стовпців
    ws.append(['brand', 'category', 'category_id', 'created', 'created_by', 'created_by_id', 'description', 'id', 'image', 'in_stock', 'is_active', 'price', 'slug', 'title', 'updated'])

    # Додати дані з моделі до файлу Excel
    for product in products:
        # Конвертувати дати та часи до місцевого часового поясу та встановити tzinfo=None
        local_created = timezone.localtime(product['created']).replace(tzinfo=None)
        local_updated = timezone.localtime(product['updated']).replace(tzinfo=None)
        ws.append([product['brand'],product['category'], product['category_id'],local_created,product['created_by'],product['created_by_id'],product['description'],product['id'],product['image'],product['in_stock'], product['is_active'],product['price'], product['slug'],product['title'], local_updated])


    # Повернути файл Excel як HTTP-відповідь
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=export.xlsx'
    wb.save(response)
    return response


def all_products(request):
    products = Product.objects.all()
    return render(request, 'store/home.html', {'products': products})

def product_detail(request, slug):
    product = get_object_or_404(Product, slug=slug, in_stock=True)
    return render(request, 'store/products/detail.html', {'product': product})

def category_list(request, category_slug):
    category = get_object_or_404(Category, slug=category_slug)
    products = Product.objects.filter(category=category)
    return render(request, 'store/products/category.html', {'category': category, 'products': products})

from django.db.models import Count
from django.shortcuts import render
from .models import Product, Category


def diagrams(request):
    categories = Category.objects.annotate(product_count=Count('product'))
    
    productsCatgId_list = []
    number_list = []
    
    for category in categories:
        productsCatgId_list.append(category.name)
        number_list.append(category.product_count)
    
    brand_dict = {}

    for product in Product.objects.all():
        brand = product.brand

        # increment brand count
        if brand in brand_dict:
            brand_dict[brand] += 1
        else:
            brand_dict[brand] = 1

    brand_list = list(brand_dict.keys())
    brand_number = list(brand_dict.values())

    context = {
        'productsCatgId_list': productsCatgId_list,
        'number_list': number_list,
        'brand_list': brand_list,
        'brand_number': brand_number
    }
    
    return render(request, 'store/diagrams.html', context)


